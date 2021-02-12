import openpyxl
import pickle


def filename_from_id(symid):
    filename = symid.zfill(5) + ".png"
    return filename


def process_one(fname, symdict, imgiddict, labelcols):
    wb = openpyxl.load_workbook(fname)
    ws = wb["Sheet1"]
    for xrow in range(3, ws.max_row + 1):
        symid = ws.cell(row=xrow, column=2).value
        symfname = filename_from_id(str(symid))
        imgiddict[symfname] = symid
        for labelcol in labelcols:
            labels = ws.cell(row=xrow, column=labelcol).value
            if labels:
                labels = str(labels)
                labels = labels.split("|")
                for label in labels:
                    label = label.strip()
                    if label not in symdict:
                        symdict[label] = [symfname]
                    elif symfname not in symdict[label]:
                        symdict[label].append(symfname)


def make_pcs_map(sdict, idict):
    process_one("Addenda Core-2004.xlsx", sdict, idict, [4, 15])
    process_one("2006 2008 2012 Addenda with Translations.xlsx", sdict, idict, [4, 15])
    process_one("PCS Classic Safeguard w Translations.xlsx", sdict, idict, [3, 4])
    process_one("PCS Supplemental 1.xlsx", sdict, idict, [3, 4])
    process_one("PCS Supplemental 2.xlsx", sdict, idict, [3, 4])
    process_one("ThinLine Translations.xlsx", sdict, idict, [3, 14])


def make_high_contrast_map(sdict, idict):
    process_one("PCS High Contrast with Translations.xlsx", sdict, idict, [3, 15])


def fix_symbol_map(fname, oname):
    with open(fname, 'rb') as fp:
        sdict = pickle.load(fp)

    newdict = {key: [value] for (key, value) in sdict.items()}

    with open(oname, 'wb') as fp:
        pickle.dump(newdict, fp)


symdict = dict()
imgiddict = dict()
fix_symbol_map("c:\\PycharmProjects\\Playpen\\symbols\\PRC\\symbol-old.map", "c:\\PycharmProjects\\Playpen\\symbols\\PRC\\symbol.map")
fix_symbol_map("c:\\PycharmProjects\\Playpen\\symbols\\ARASAAC BW\\symbol-old.map", "c:\\PycharmProjects\\Playpen\\symbols\\ARASAAC BW\\symbol.map")
fix_symbol_map("c:\\PycharmProjects\\Playpen\\symbols\\ARASAAC Color\\symbol-old.map", "c:\\PycharmProjects\\Playpen\\symbols\\ARASAAC Color\\symbol.map")
# make_pcs_map(symdict, imgiddict)
# make_high_contrast_map(symdict, imgiddict)
# with open("symbol.map", 'wb') as f:
#     pickle.dump(symdict, f)
# with open("img-id.map", 'wb') as f:
#     pickle.dump(imgiddict, f)
